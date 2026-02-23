import os

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

ref_dir = os.getenv('ref_dir')
salary_ledger = os.getenv('salary_ledger')

def process_insurance(file_prefix, file_ext, extract_cols, col_mapping):
    # 디렉토리 내 파일 검색
    candidates = [f for f in os.listdir(ref_dir) if f.startswith(file_prefix) and f.endswith(file_ext)]

    if not candidates:
        print(f"처리할 파일({file_prefix}*{file_ext})을 찾지 못했습니다.")
        return

    if len(candidates) > 1:
        print("여러 파일이 발견되었습니다. 선택해주세요:")
        for idx, fname in enumerate(candidates):
            print(f"{idx + 1}. {fname}")
        try:
            choice = int(input("번호 입력: ")) - 1
            filename = candidates[choice]
        except (ValueError, IndexError):
            print("잘못된 입력입니다.")
            return
    else:
        filename = candidates[0]

    file_path = os.path.join(ref_dir, filename)
    # 엑셀 파일 읽기 및 특정 컬럼 추출
    # 1. 헤더 위치 찾기 ('성명'이 있는 행을 헤더로 간주)
    df_temp = pd.read_excel(file_path, header=None)
    header_idx = -1
    for idx, row in df_temp.iterrows():
        if '성명' in [str(v).strip() for v in row.values]:
            header_idx = idx
            break
    
    if header_idx == -1:
        print(f"'{filename}' 파일에서 '성명' 헤더를 찾을 수 없습니다.")
        return

    df = pd.read_excel(file_path, header=header_idx)

    # 2. 컬럼명 매칭 (공백, 줄바꿈, _ 제거하여 유연하게 매칭)
    df_cols_clean = {str(c).replace(' ', '').replace('\n', '').replace('_', ''): c for c in df.columns}
    
    try:
        real_cols = [df_cols_clean[col.replace(' ', '').replace('\n', '').replace('_', '')] for col in extract_cols]
    except KeyError as e:
        print(f"컬럼 매칭 실패: {e}")
        print(f"파일 내 컬럼 목록: {list(df.columns)}")
        return

    source_df = df[real_cols].rename(columns=dict(zip(real_cols, extract_cols)))
    print(f"소스 파일 로드: {filename}")

    # 타겟 엑셀 파일 열기 및 업데이트
    try:
        wb = load_workbook(salary_ledger)
        if '사대보험' not in wb.sheetnames:
            print("'사대보험' 시트를 찾을 수 없습니다.")
            return
        ws = wb['사대보험']

        # 헤더 찾기
        header_map = {}
        header_row_idx = None
        for r_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
            if '성명' in row:
                header_row_idx = r_idx
                target_cols = ['성명'] + list(col_mapping.values())
                for c_idx, val in enumerate(row):
                    if val in target_cols:
                        header_map[val] = c_idx # 0-based index
                break
        
        if not header_map or '성명' not in header_map:
            print("필요한 헤더(성명)를 찾을 수 없습니다.")
            return

        # 데이터 매핑 및 입력
        source_data = source_df.set_index('성명').to_dict('index')
        
        count = 0
        for row in ws.iter_rows(min_row=header_row_idx + 1):
            name_cell = row[header_map['성명']]
            name = name_cell.value
            
            if name in source_data:
                for src_col, target_col in col_mapping.items():
                    if target_col in header_map:
                        val = source_data[name].get(src_col)
                        if pd.notna(val):
                            row[header_map[target_col]].value = int(val)
                count += 1
        
        wb.save(salary_ledger)
        print(f"업데이트 완료: {count}건 처리됨. ({salary_ledger})")

    except Exception as e:
        print(f"오류 발생: {e}")

if __name__ == '__main__':
    '''
    국민연금>결정내역>결정내역통보서(2차)>통합저장 (매월20일 경)
    건강보험>받은문서함>가입자고지내역서>파일변환 (매월20일경)
    '''
    print("작업을 선택하세요:")
    print("1. 건강보험 (보험료_고지(산출)_내역서)")
    print("2. 국민연금 (2차결정내역통보서)")
    choice = input("선택: ")

    if choice == '1':
        process_insurance(
            file_prefix='보험료_고지(산출)_내역서',
            file_ext='.xls',
            extract_cols=['성명', '고지금액', '요양고지보험료'],
            col_mapping={'고지금액': '건강', '요양고지보험료': '요양'}
        )
    elif choice == '2':
        process_insurance(
            file_prefix='2차결정내역통보서',
            file_ext='.xlsx',
            extract_cols=['성명', '총부담금계_(본인기여금)(원)', '국고지원금액_본인기여금(원)'],
            col_mapping={'총부담금계_(본인기여금)(원)': '연금', '국고지원금액_본인기여금(원)': '두루누리'}
        )
    else:
        print("잘못된 선택입니다.")
