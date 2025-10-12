import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
from io import BytesIO

# 엑셀 파일 불러오기
file_path = './data/원장.xlsx'
df = pd.read_excel(file_path)

# 데이터 전처리
df['날짜'] = pd.to_datetime(df['날짜']).dt.date
df['월'] = pd.to_datetime(df['날짜']).dt.month
df['차변'] = df['차변'].fillna(0)
df['대변'] = df['대변'].fillna(0)

# 현장명 정리 (동일 현장 통합)
df['현장명'] = df['현장명'].str.split('(').str[0]


# 1. 현장별 수익/비용 정리
def create_site_summary_report():
    # 수익 데이터 필터링
    revenue_df = df[df['구분'] == '수익'].copy()
    revenue_df['순수익'] = revenue_df['대변'] - revenue_df['차변']

    # 비용 데이터 필터링
    expense_df = df[df['구분'] == '비용'].copy()
    expense_df['비용'] = expense_df['차변'] - expense_df['대변']

    # 가지급 데이터 필터링
    temp_adv_df = df[(df['계정과목'] == '가지급') & (df['담당'] == df['거래처'])].copy()
    temp_adv_df['가지급'] = temp_adv_df['대변'] - temp_adv_df['차변']

    # 현장별 수익 집계
    site_revenue = revenue_df.groupby('현장명')['순수익'].sum().reset_index()
    site_revenue.columns = ['현장명', '총수익']

    # 현장별 비용 집계
    site_expense = expense_df.groupby('현장명')['비용'].sum().reset_index()
    site_expense.columns = ['현장명', '총비용']

    # 가지급 현장별, 집계
    temp_adv = temp_adv_df.groupby('현장명')['가지급'].sum().reset_index()
    temp_adv.columns = ['현장명', '가지급계']

    # 현장별 가지급정리
    site_temp = pd.merge(temp_adv, site_expense, on='현장명', how='outer').fillna(0)
    site_temp['운영비현황'] = site_temp['가지급계'] - site_temp['총비용']
    site_temp = site_temp.sort_values('현장명', ascending=False)

    # 현장별 순이익 계산
    site_summary = pd.merge(site_revenue, site_expense, on='현장명', how='outer').fillna(0)
    site_summary['순이익'] = site_summary['총수익'] - site_summary['총비용']
    site_summary['수익률'] = (site_summary['순이익'] / site_summary['총수익'] * 100).round(2)
    site_summary = site_summary.sort_values('순이익', ascending=False)

    # 월별 추이 분석
    monthly_revenue = revenue_df.groupby(['현장명', '월'])['순수익'].sum().unstack().fillna(0)
    monthly_expense = expense_df.groupby(['현장명', '월'])['비용'].sum().unstack().fillna(0)

    # 엑셀 파일 생성
    with pd.ExcelWriter('현장별_수익_정리.xlsx') as writer:
        # 현장별 요약
        site_summary.to_excel(writer, sheet_name='현장별 요약', index=False)

        # 현장별 운영비 요약
        site_temp.to_excel(writer, sheet_name='현장별 운영비', index=False)

        # 월별 수익 추이
        monthly_revenue.to_excel(writer, sheet_name='월별 수익 추이')

        # 월별 비용 추이
        monthly_expense.to_excel(writer, sheet_name='월별 비용 추이')

        # 원본 데이터 (필터링된)
        pd.concat([revenue_df, expense_df]).sort_values('날짜').to_excel(
            writer, sheet_name='원본 데이터', index=False)

    print("현장별 수익 정리 보고서가 생성되었습니다: 현장별_수익_정리.xlsx")


# 2. 담당자 평가 보고서
def create_employee_evaluation_report():
    # 담당자별 현장 데이터 추출
    employee_df = df[df['담당'].notna()].copy()

    # 담당자별 현장 수
    site_count = employee_df.groupby('담당')['현장명'].nunique().reset_index()
    site_count.columns = ['담당자', '담당 현장 수']

    # 담당자별 처리 건수
    transaction_count = employee_df.groupby('담당')['코드'].count().reset_index()
    transaction_count.columns = ['담당자', '처리 건수']

    # 담당자별 총액 처리량 (차변 + 대변)
    amount_processed = employee_df.groupby('담당')[['차변', '대변']].sum().reset_index()
    amount_processed['총처리액'] = amount_processed['차변'] + amount_processed['대변']
    amount_processed = amount_processed[['담당', '총처리액']]
    amount_processed.columns = ['담당자', '총처리액']

    # 담당자별 수익/비용 분석
    revenue_expense = employee_df[employee_df['구분'].isin(['수익', '비용'])].copy()
    revenue_expense['금액'] = revenue_expense.apply(
        lambda x: x['대변'] if x['구분'] == '수익' else -x['차변'], axis=1)

    employee_performance = revenue_expense.groupby('담당')['금액'].agg(['sum', 'mean', 'count']).reset_index()
    employee_performance.columns = ['담당자', '순수익', '평균거래액', '거래건수']

    # 모든 지표 병합
    evaluation_report = site_count.merge(
        transaction_count, on='담당자').merge(
        amount_processed, on='담당자').merge(
        employee_performance, on='담당자')

    # 평가 점수 계산 (간단한 가중치 적용)
    evaluation_report['담당현장점수'] = evaluation_report['담당 현장 수'] / evaluation_report['담당 현장 수'].max() * 100
    evaluation_report['처리건수점수'] = evaluation_report['처리 건수'] / evaluation_report['처리 건수'].max() * 100
    evaluation_report['처리액점수'] = evaluation_report['총처리액'] / evaluation_report['총처리액'].max() * 100
    evaluation_report['수익점수'] = (evaluation_report['순수익'] - evaluation_report['순수익'].min()) / (
            evaluation_report['순수익'].max() - evaluation_report['순수익'].min()) * 100

    evaluation_report['종합점수'] = (
            evaluation_report['담당현장점수'] * 0.2 +
            evaluation_report['처리건수점수'] * 0.2 +
            evaluation_report['처리액점수'] * 0.3 +
            evaluation_report['수익점수'] * 0.3
    ).round(1)

    evaluation_report = evaluation_report.sort_values('종합점수', ascending=False)

    # 엑셀 파일 생성 (서식 적용)
    wb = Workbook()
    ws = wb.active
    ws.title = "담당자 평가"

    # 헤더 작성
    headers = ['담당자', '담당 현장 수', '처리 건수', '총처리액',
               '순수익', '평균거래액', '거래건수', '종합점수']
    ws.append(headers)

    # 데이터 작성
    for _, row in evaluation_report.iterrows():
        ws.append([
            row['담당자'],
            row['담당 현장 수'],
            row['처리 건수'],
            row['총처리액'],
            row['순수익'],
            row['평균거래액'],
            row['거래건수'],
            row['종합점수']
        ])

    # 서식 적용
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            if cell.column_identifier in ['D', 'E']:  # 금액 컬럼
                cell.number_format = '#,##0'

    # 컬럼 너비 조정
    col_widths = [15, 12, 12, 15, 15, 15, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # 차트 생성
    fig, ax = plt.subplots(figsize=(10, 6))
    evaluation_report.plot(x='담당자', y='종합점수', kind='bar', ax=ax, color='skyblue')
    ax.set_title('담당자 종합 평가 점수')
    ax.set_ylabel('점수')
    plt.tight_layout()

    # 차트를 이미지로 저장하고 엑셀에 삽입
    chart_image = BytesIO()
    plt.savefig(chart_image, format='png')
    plt.close()

    from openpyxl.drawing.image import Image
    img = Image(chart_image)
    img.anchor = 'I2'
    ws.add_image(img)

    # 파일 저장
    wb.save('담당자_평가_보고서.xlsx')
    print("담당자 평가 보고서가 생성되었습니다: 담당자_평가_보고서.xlsx")


# 보고서 생성 실행
create_site_summary_report()
create_employee_evaluation_report()