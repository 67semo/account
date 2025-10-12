import openpyxl
import os
from dotenv import load_dotenv

load_dotenv()

data_dr = os.getenv('data_dir')
format_dir = os.getenv('format')
form_nm = '보고서.xlsx'
sheet_nm = '보통예금'

# --- 1. 파일 경로 및 이름 설정 ---
# 원본 양식 파일 이름
source_filename = os.path.join(format_dir, form_nm)
# 저장할 새 보고서 파일 이름
report_filename = os.path.join(data_dr,'report.xlsx')

# --- 2. 입력할 데이터 정의 ---
# 딕셔너리 형태로, '셀 주소': '입력할 값'을 정의합니다.
data_to_input = {
    'B2': '2025-10-04',          # 보고서 날짜
    'C4': '김철수',              # 담당자 이름
    'D5': 1500,                  # 숫자 데이터 (예: 판매 수량)
    'E6': '=D5*1000'             # 엑셀 수식도 입력 가능합니다.
}

# --- 3. 파일 열기 및 데이터 입력 ---
try:
    # 1. 원본 엑셀 파일 열기 (Workbook 로드)
    workbook = openpyxl.load_workbook(source_filename)

    # 2. 지정된 시트(worksheet) 선택
    if sheet_nm in workbook.sheetnames:
        sheet = workbook[sheet_nm]
    else:
        print(f"ERROR: 시트 '{sheet_nm}'을(를) 찾을 수 없습니다. (시트 목록: {workbook.sheetnames})")
        exit()

    # 3. 데이터 입력
    print(f"'{source_filename}' 파일의 '{sheet_nm}' 시트에 데이터를 입력합니다.")
    for cell_address, value in data_to_input.items():
        sheet[cell_address] = value
        print(f"  -> {cell_address}: {value} 입력 완료")

    # --Delete unnecessary sheets--
    sheets_to_delete = [s for s in workbook.sheetnames if s != sheet_nm]
    for s in sheets_to_delete:
        std = workbook[s]
        workbook.remove(std)
    print(f"불필요한 시트 {sheets_to_delete} 삭제 완료.")

    # --- 4. 새 파일로 저장 ---
    workbook.save(report_filename)
    print(f"\nSUCCESS: 데이터를 입력하고 '{report_filename}' 파일로 저장했습니다.")

except FileNotFoundError:
    print(f"ERROR: 원본 파일 '{source_filename}'을(를) 현재 디렉토리에서 찾을 수 없습니다.")
except Exception as e:
    print(f"처리 중 예상치 못한 에러 발생: {e}")