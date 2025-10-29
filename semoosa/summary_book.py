import pandas as pd
from openpyxl import load_workbook
import math, os
from dotenv import load_dotenv

load_dotenv()

data_dir = os.getenv('data_dir')
book_file = '25장부.xlsx'
book_path = os.path.join(data_dir, book_file)
book_sheet = "25년장부"
format_dir = os.getenv('project') + '/form'  # 포맷파일이 있는 폴더
report = os.path.join(format_dir, '보고서.xlsx')

def read_data():    # 장부읽어와 데이터 프레임으로 리턴
    try:
        df = pd.read_excel(book_path, sheet_name=book_sheet, header=3)

    except FileNotFoundError:
        print(f"Error : {book_path} 파일이 존재하지 않습니다.")
    except Exception as e:
        print(f"{e} 에러 발생")
    return df

# 보통예금 잔고정산때 사용
def current_account(df, dt):    # 보통예금만을 추리고 계정별로 함산후 잔액을 딕려서리로 리턴
    df['날짜'] = pd.to_datetime(df['날짜'])
    bank_books = df[(df['계정과목'] == '보통예금') & (df['날짜'].dt.date <= dt)]       # 보통예금 통장
    books_lst = bank_books['거래처'].unique().tolist()
    #print(books_lst)
    balence_dic = {}
    for book in books_lst:
        book_df = bank_books[bank_books['거래처'] == book]
        #balence = book_df['차변'].sum() - book_df['대변'].sum()
        balence = debit_credit_valence(book_df)
        balence_dic[book] = int(balence)
    return balence_dic

# 보통예금 잔고정산때 사용
def write_workbook(data):   # 보고서에 기입
    print(report)
    workbook = load_workbook(report)
    sheet = workbook['보통예금']
    # 데이터를 입력할 시작 셀 (예: B5 셀)
    start_row = 5
    start_col = 2  # B열
    index = 0

    # DataFrame의 행들을 순회하며 값 입력
    for k, v in data.items():
        print(k, v)
        # 엑셀 셀에 값 입력
        sheet.cell(row=start_row + index, column=start_col, value=k)
        sheet.cell(row=start_row + index, column=start_col + 1, value=v)
        index += 1

    # 변경사항 저장
    report_last = os.path.join(data_dir, '잔액보고.xlsx')
    workbook.save(report_last )
    print("성공적으로 for 루프를 사용하여 값을 입력했습니다.")

def report_for_director(name, dic_data):
    # 템플릿 파일 열기
    form1 = os.path.join(format_dir, '보고서.xlsx')
    wb = load_workbook(form1)

    # 원본 시트 복사
    src_sheet = wb['이사정산']
    input_lines = [6, 7, 11, 15, 16]
    fist_item = True
    for key, lst in dic_data.items():
        if fist_item == True:
            active_sheet = src_sheet
            fist_item = False
        else:
            active_sheet = wb.copy_worksheet(src_sheet)
        active_sheet.title = key[:5]  # 새 시트명 지정
        for i in range(len(input_lines)):
            row = input_lines[i]
            #print(lst[2*i], active_sheet.cell(row,3).value)
            active_sheet.cell(row, 4, lst[2*i])
            active_sheet.cell(row, 5, lst[2*i+1])

    # 저장 (기존 템플릿 파일에 적용)
    f_name = name + '.xlsx'
    f_path = os.path.join(data_dir, f_name)
    wb.save(f_path)
    print(f"'{name}' 정산 자료가 저장 되었습니다.")

def debit_credit_valence(df):
    return int(df['대변'].sum() - df['차변'].sum())

def freelancer(df, frate=0.07):  # 프리랜서 현정정리
    df1 = df[(df['구분'] == '부채') & (df['계정과목'] == '선수금')]       # 선수금
    debt_bal = debit_credit_valence(df1)        # 대변 - 차변
    if debt_bal:                        # 선수금이 있을경우
        df1 = df[(df['구분'] == '부채') & (df['계정과목'] == '부가세예수금')]   # 부가세예수금
        debt_surtax = debit_credit_valence(df1)     # 대변 - 차변
        if debt_bal * 0.1 != debt_surtax:           # 부가세예수금과 선수금의 비율이 맞지않으면 선수금기준으로 부가세예수금처리 --> 이해하지 못할 처리??
            debt_surtax = debt_bal * 0.1
    else:
        debt_surtax = 0                       # 선수금이 없으면, 부가세예수금은 0원

    df1 = df[df['구분'] == '수익']             # 수익처리 
    revenue = debit_credit_valence(df1)
    if revenue:
        df1 = df[(df['구분'] == '부채') & (df['계정과목'] == '부가세예수금')]   # 부가세예수금과 선수금의 비율이 맞지않으면 선수금기준으로 부가세예수금처리 --> 이해하지 못할 처리??
        rev_surtax = debit_credit_valence(df1)
        if revenue * 0.1 != rev_surtax:      # 부가세예수금과 수익의 비율이 맞지않으면 선수금기준으로 부가세예수금처리 --> 이해하지 못할 처리??
            rev_surtax = revenue * 0.1
    else:
        rev_surtax = 0

    df1 = df[(df['계정과목'] == '가지급') & (~df['거래처'].isin(['고영수', '대표이사']) )]  # 대표이사를 제외한 가지급 (담당이사의 가지급)
    delector_deposit = debit_credit_valence(df1)
    df1 = df[df['구분'] == '비용']           # 비용처리 
    expense = - debit_credit_valence(df1)   
    df1 = df[df['계정과목'] == '부가세대급금']  # 부가세대급금
    exp_surtax = - debit_credit_valence(df1)

    head_fee = math.ceil((debt_bal+revenue) * frate/10)*10  # 본사 소득
    # 부채(선금), 부채부가세, 수익, 수익부가세, 담당이사 가지급, 0, 비용, 비용부가세, 본사소득, 0
    items = [debt_bal, debt_surtax, revenue, rev_surtax,delector_deposit, 0, expense, exp_surtax, head_fee,0]
    return items

def account_for_directors(ttl_df):
    # sites = account_df[(account_df['담당'] != '') & (~account_df['담당'].str.startswith('고', na=False))]
    sites = ttl_df.dropna(subset='담당')
    # sites = sites[~sites['담당'].str.startswith('고', na=False)]
    sites = sites[~sites['담당'].str.startswith('고')]

    stat = sites['현장명'].unique()

    grouped_sites = sites.groupby('담당')
    manager = {}
    for name, df in grouped_sites:
        sits = df.groupby('현장명')
        rt_dic = {}
        for sname, sdf in sits:
            if sname == '복합커뮤니티센터 건립중설비':
                fee_rate = 0.05
            else:
                fee_rate = 0.07
            site_rt = freelancer(sdf, fee_rate)
            rt_dic[sname] = site_rt
        report_for_director(name, rt_dic)
    print('이사들의 정산이 마무리 되었습니다.')

if __name__ == '__main__':
    account_df = read_data()
    date_str = input('날자를 입력하세요 (m/d)')

    from datetime import date

    if date_str == '':
        date_rt = date.today()
    else:
        cur_year = date.today().year
        mon, day = map(int, date_str.split('/'))
        date_rt = date(cur_year, mon, day)

    account_for_directors(account_df)   # 이사별정산하여 개별엑셀파일을 data 디렉토리에 저장
    deposit_dic = current_account(account_df, date_rt)  #보통예금 정리
    write_workbook(deposit_dic)  # 보통예금정리하여 잔액보고서에 기록
  