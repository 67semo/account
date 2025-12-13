# 카드사용내역중 부가세 처리하여 세무사 양식에 저장
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
import os

load_dotenv()

filltering_lst = ['재료비','폐기물처리비', '차량유지비', '소모재료비','소모품비','공구비']
require_col = ['카드번호', '승인일자', '승인금액(원화)', '거래금액(원화)', '부가세', '회계코드명', '가맹점사업자번호', '가맹점명']

# Excel 파일에서 특정 열의 값들을 리스트화하여 반환
def extract_req_col_from_excel(file_path, require_col_column_name, sheet_name=0):
    try:
        # Excel 파일을 데이터프레임으로 읽기
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        # print(df)
        if require_col_column_name not in df.columns:
            print(f"오류: '{require_col_column_name}' 열을 찾을 수 없습니다. 사용 가능한 열: {df.columns.tolist()}")
            return []

        # 해당 열의 고유한 값들을 리스트로 추출
        resp_lst = df[require_col_column_name].dropna().unique().tolist()
        return resp_lst
    except FileNotFoundError:
        print(f"오류: 파일을 찾을 수 없습니다 - {file_path}")
        return []
    except Exception as e:
        print(f"데이터 처리 중 오류 발생: {e}")
        return []

def analize_book(df, col_name):     # 회계코드(실질은 계정과목)의 값이 filltering_lst에 있는 값이거나 부가세가 있는경우
    vat_filter_mask = (df[col_name].isin(filltering_lst)) & (df['부가세'] != 0)
    vat_applicable_rows = df[vat_filter_mask].copy()
    result_df = vat_applicable_rows[require_col]
    result_df = result_df.sort_values('승인일자', ascending=True)
    return result_df

def write_to_excel(df, file_path, sheet_nm, st_row, st_col):
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(
                writer,
                sheet_name=sheet_nm,
                startrow=st_row,
                startcol=st_col,
                index=False,
                header=False
            )
        print(f"성공: '{file_path}' 파일의 '{sheet_nm}' 시트에 데이터가 성공적으로 입력되었습니다.")
        return True

    except FileNotFoundError:
        print(f"오류: Excel 파일을 찾을 수 없습니다 - {file_path}")
        return False
    except ValueError as e:
        print(f"오류: Excel 쓰기 중 문제 발생 - {e}. 시트 이름 또는 시작 셀을 확인하세요.")
        return False
    except Exception as e:
        print(f"예상치 못한 오류 발생: {e}")
        return False

def modify_for_num(num_str):
    if not isinstance(num_str, str):
        num_str = str(num_str)
    import re
    cleaned_st = re.sub(r'\s+', '', num_str)
    cleaned_st = cleaned_st.replace(',', '')
    return cleaned_st

def clean_col_data(df):
    # 숫자에서 전단의 기호와 천구분단위 삭제
    cols = require_col[2:5]     # 금액에 관련된 컬럼들
    for col in cols:
        df[col] = df[col].apply(modify_for_num)
        df[col] = pd.to_numeric(df[col], errors="coerce")
    #print(df[cols])
    return df

def for_semusa_form(filterd_df):  # 세무사양식을 체우기위한 과장
    column_name = '회계코드'
    data_dir = os.getenv('data_dir')
    output_file = os.path.join(data_dir, '세무사양식.xlsx')
    sheet_nm = '카드매입'
    wb = load_workbook(output_file)
    ws = wb[sheet_nm]
    srow = ws.max_row
    scol = 0
    result_df = analize_book(filterd_df, column_name)

    write_to_excel(result_df, output_file, sheet_nm, srow, scol)


if __name__ == '__main__':
    # --- 사용 예시 ---
    data_dir = os.getenv('data_dir')
    output_file = os.path.join(data_dir, '세무사양식.xlsx')
    input_file = os.path.join(data_dir, '카드자료.xlsx')   # './data/카드자료.xlsx' # 실제 Excel 파일 경로로 변경하세요
    column_name = '회계코드'
    # output_file = './data/세무사양식.xlsx'
    sheet_nm = '카드매입'
    wb = load_workbook(output_file)
    ws = wb[sheet_nm]
    srow = ws.max_row
    scol = 0

    df = pd.read_excel(input_file, sheet_name=0)
    df = clean_col_data(df)
    df = analize_book(df, column_name)

    write_to_excel(df, output_file, sheet_nm, srow, scol)

    '''
    *. Excel 파일에서 특정 열의 값들을 리스트화하여 반환 실행
    result_llst = extract_req_col_from_excel(excel_file, column_name)
    print(f"Excel 파일에서 추출된 회계 코드: {result_llst}")
    '''
